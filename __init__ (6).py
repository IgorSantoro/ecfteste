"""Reclassificação de códigos do e-Lalur / e-Lacs (M300 e M350).

Transfere valores entre códigos, período a período de apuração (``M030``:
``A00`` anual, ``A01``–``A12`` balancetes acumulados, ``T01``–``T04``
trimestres), com base na planilha De→Para (colunas Janeiro…Dezembro, Anual).

Como o lançamento de origem é localizado, em cada período:

1. Pelo **código De** e, quando o lançamento possui filhos de vínculo, pela
   conta indicada na planilha — ``PARTE B`` casa o filho **M305/M355**
   (``COD_CTA_B`` = coluna G); número de conta casa o filho **M310/M360**
   (``COD_CTA`` = coluna F). Contas são comparadas ignorando máscara
   (pontos, hífens, espaços).
2. Se o lançamento **não tem nenhum filho de vínculo** (situação comum: a
   ECF traz o valor agregado do código, e a decomposição por conta é
   controle contábil externo), o casamento é **pelo código**. Nesse caso,
   várias linhas da planilha podem sacar, cada uma, sua parcela do mesmo
   lançamento — a soma das parcelas não pode exceder o valor.
3. Havendo mais de um lançamento candidato no mesmo período, o desempate é
   pela **descrição** (e pelo valor); sem distinção suficiente → conflito.

Cada saque vira uma operação: subtrai do lançamento De e cria um lançamento
com o código **Para** (descrição oficial da tabela de códigos), levando o
vínculo quando ele existe (conta movida; parcela da Parte B dividida). Se o
De zerar, é removido. Montante igual ao valor do lançamento → o próprio
lançamento muda de código (troca integral).

Nada é aplicado sem prévia; regras sem alvo recebem um **diagnóstico**
explicando o que existe no arquivo e por que não casou.
"""
from __future__ import annotations

import json
import os
import re
import unicodedata
from dataclasses import dataclass, field
from difflib import SequenceMatcher
from typing import Dict, List, Optional, Sequence, Tuple

from ..models import ArquivoECF, Registro

# --------------------------------------------------------------------------- #
# CONSTANTES
# --------------------------------------------------------------------------- #
PARES = {
    "M300": ("M305", "M310"),   # e-Lalur  (IRPJ)
    "M350": ("M355", "M360"),   # e-Lacs   (CSLL)
}
FILHOS = {
    "M300": ("M305", "M310", "M312", "M315"),
    "M350": ("M355", "M360", "M362", "M365"),
}
CAMPO_CODIGO = 1
CAMPO_DESC = 2
CAMPO_TIPO = 3
CAMPO_REL = 4
CAMPO_VALOR = 5
CAMPO_CONTA = 1
CAMPO_VL_CTA = 2
CAMPO_IND = 3

RE_PARTE_B = re.compile(r"parte\s*b", re.IGNORECASE)

COLUNAS_PADRAO = {
    "codigo": "B",
    "descricao": "C",
    "relacionamento": "F",
    "conta_b": "G",
    "imposto": "I",
    "meses_inicio": "K",
    "valor": "W",
    "de": "Y",
    "para": "Z",
}

PERIODO_COLUNA = {**{f"A{m:02d}": m - 1 for m in range(1, 13)}, "A00": 12,
                  "T01": 2, "T02": 5, "T03": 8, "T04": 11}

LIMIAR_DESEMPATE = 0.45
TOL = 0.02


# --------------------------------------------------------------------------- #
# Utilidades
# --------------------------------------------------------------------------- #
def normalizar(texto: str) -> str:
    if not texto:
        return ""
    t = unicodedata.normalize("NFKD", str(texto))
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = re.sub(r"[^A-Za-z0-9 ]+", " ", t).upper()
    return re.sub(r"\s+", " ", t).strip()


def similaridade(a: str, b: str) -> float:
    na, nb = normalizar(a), normalizar(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    return SequenceMatcher(None, na, nb).ratio()


def normaliza_conta(c: str) -> str:
    """Compara contas ignorando máscara: '6.105.21-99' == '61052199'."""
    return re.sub(r"[^0-9A-Za-z]", "", c or "").upper()


def parse_valor_ecf(s: str) -> Optional[float]:
    s = (s or "").strip()
    if not s:
        return None
    try:
        return float(s.replace(".", "").replace(",", "."))
    except ValueError:
        return None


def fmt_valor_ecf(v: float) -> str:
    return f"{abs(v):.2f}".replace(".", ",")


def fmt_brl(v: float) -> str:
    inteiro, dec = f"{abs(v):,.2f}".split(".")
    return inteiro.replace(",", ".") + "," + dec


class TabelaCodigos:
    """Tabela oficial código→descrição/tipo de M300 e M350."""

    def __init__(self, pasta_layouts: Optional[str] = None) -> None:
        pasta = pasta_layouts or os.path.join(
            os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
            "layouts")
        self._dados: Dict[str, Dict[str, dict]] = {}
        caminho = os.path.join(pasta, "codigos_ecf_0012.json")
        if os.path.exists(caminho):
            with open(caminho, encoding="utf-8") as f:
                self._dados = json.load(f)

    @property
    def disponivel(self) -> bool:
        return bool(self._dados)

    def existe(self, pai: str, codigo: str) -> bool:
        return codigo in self._dados.get(pai, {})

    def descricao(self, pai: str, codigo: str) -> str:
        return self._dados.get(pai, {}).get(codigo, {}).get("descricao", "")


# --------------------------------------------------------------------------- #
# Estruturas
# --------------------------------------------------------------------------- #
@dataclass
class RegraReclass:
    linha: int
    de: str
    para: str
    relacionamento: str = ""
    conta_b: str = ""
    descricao: str = ""
    imposto: str = "AMBOS"
    valores: List[Optional[float]] = field(default_factory=lambda: [None] * 13)

    @property
    def por_parte_b(self) -> bool:
        return bool(RE_PARTE_B.search(self.relacionamento or ""))

    @property
    def valor(self) -> Optional[float]:
        return self.valores[12]

    @valor.setter
    def valor(self, v: Optional[float]) -> None:
        self.valores[12] = v

    @property
    def tem_valores(self) -> bool:
        return any(v is not None and abs(v) > TOL for v in self.valores)

    def montante(self, periodo: str) -> Optional[float]:
        i = PERIODO_COLUNA.get(periodo)
        if i is None:
            return None
        v = self.valores[i]
        return None if v is None or abs(v) <= TOL else abs(v)

    def aplica_em(self, pai: str) -> bool:
        imp = (self.imposto or "AMBOS").upper()
        if "IRPJ" in imp and "CSLL" not in imp and "AMBOS" not in imp:
            return pai == "M300"
        if "CSLL" in imp and "IRPJ" not in imp and "AMBOS" not in imp:
            return pai == "M350"
        return True

    @property
    def chave(self) -> str:
        alvo = self.conta_b if self.por_parte_b else self.relacionamento
        tipo = "Parte B" if self.por_parte_b else "conta"
        return f"{self.de} · {tipo} {alvo}"


@dataclass
class Correspondencia:
    regra: RegraReclass
    indice: int
    registro: str
    periodo: str
    montante: float
    modo: str                     # 'integral' | 'parcial' | 'soma_destino'
    antes: str                    # linha De antes desta operação
    depois: str                   # linha De após ('' quando removido)
    novo: str = ""                # lançamento criado (destino inexistente)
    destino_indice: Optional[int] = None    # destino já existente
    destino_antes: str = ""
    destino_depois: str = ""
    conta_casada: str = ""
    criterio: str = "vínculo"
    score: float = 1.0


@dataclass
class ResultadoReclass:
    correspondencias: List[Correspondencia] = field(default_factory=list)
    sem_alvo: List[Tuple[RegraReclass, str]] = field(default_factory=list)
    conflitos: List[str] = field(default_factory=list)
    problemas: List[str] = field(default_factory=list)
    avisos: List[str] = field(default_factory=list)

    @property
    def aplicavel(self) -> bool:
        return bool(self.correspondencias) and not self.conflitos


# --------------------------------------------------------------------------- #
# Leitura da planilha
# --------------------------------------------------------------------------- #
class LeitorReclass:
    def __init__(self, colunas: Optional[Dict[str, str]] = None) -> None:
        self.colunas = dict(COLUNAS_PADRAO)
        if colunas:
            self.colunas.update({k: v for k, v in colunas.items() if v})

    @staticmethod
    def _idx(letra: str) -> int:
        n = 0
        for ch in letra.strip().upper():
            if ch.isalpha():
                n = n * 26 + (ord(ch) - 64)
        return n - 1

    @staticmethod
    def _txt(v) -> str:
        if v is None:
            return ""
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v).strip()

    def ler(self, caminho_ou_bytes, aba: Optional[str] = None
            ) -> Tuple[List[RegraReclass], List[str]]:
        import io

        from openpyxl import load_workbook

        origem = (io.BytesIO(caminho_ou_bytes)
                  if isinstance(caminho_ou_bytes, (bytes, bytearray))
                  else caminho_ou_bytes)
        wb = load_workbook(origem, data_only=True, read_only=True)
        ws = wb[aba] if aba and aba in wb.sheetnames else wb.active

        i_de = self._idx(self.colunas["de"])
        i_para = self._idx(self.colunas["para"])
        i_rel = self._idx(self.colunas["relacionamento"])
        i_cb = self._idx(self.colunas["conta_b"])
        i_desc = self._idx(self.colunas["descricao"])
        i_imp = self._idx(self.colunas.get("imposto", "I"))
        i_mes0 = self._idx(self.colunas.get("meses_inicio", "K"))
        i_anual = self._idx(self.colunas.get("valor", "W"))

        regras: List[RegraReclass] = []
        avisos: List[str] = []

        for n, linha in enumerate(ws.iter_rows(values_only=True), start=1):
            def val(i: int) -> str:
                return self._txt(linha[i]) if 0 <= i < len(linha) else ""

            def num(i: int) -> Optional[float]:
                if not (0 <= i < len(linha)):
                    return None
                v = linha[i]
                return float(v) if isinstance(v, (int, float)) else None

            de, para = val(i_de), val(i_para)
            if not de or not para:
                continue
            if not re.match(r"^[\d.,]+$", de.replace(" ", "")):
                continue
            if de == para:
                avisos.append(f"Linha {n}: código {de} já é igual ao "
                              f"destino — ignorada.")
                continue

            valores: List[Optional[float]] = [num(i_mes0 + m)
                                              for m in range(12)]
            valores.append(num(i_anual))

            regras.append(RegraReclass(
                linha=n, de=de, para=para,
                relacionamento=val(i_rel), conta_b=val(i_cb),
                descricao=val(i_desc), imposto=val(i_imp) or "AMBOS",
                valores=valores))

        wb.close()
        return regras, avisos


# --------------------------------------------------------------------------- #
# Motor
# --------------------------------------------------------------------------- #
@dataclass
class _Candidato:
    idx: int
    pai: str
    periodo: str
    conta: str          # '' = casado pelo código (lançamento sem vínculos)
    criterio: str
    score: float


class Reclassificador:
    def __init__(self, alterar_descricao: bool = False,
                 tabela: Optional[TabelaCodigos] = None) -> None:
        self.alterar_descricao = alterar_descricao
        self.tabela = tabela or TabelaCodigos()

    # ------------------------------------------------------------------ #
    @staticmethod
    def _segmentos(regs: List[Registro]) -> List[Tuple[str, int, int]]:
        marcos = [(i, r.campo(3)) for i, r in enumerate(regs)
                  if r.reg == "M030"]
        if not marcos:
            return [("A00", 0, len(regs))]
        seg = []
        for k, (i, per) in enumerate(marcos):
            fim = marcos[k + 1][0] if k + 1 < len(marcos) else len(regs)
            seg.append((per, i, fim))
        return seg

    @staticmethod
    def _filhos(regs: List[Registro], i: int, pai: str) -> List[int]:
        j = i + 1
        out = []
        while j < len(regs) and regs[j].reg in FILHOS[pai]:
            out.append(j)
            j += 1
        return out

    def _vinculo(self, regs: List[Registro], i: int, pai: str,
                 regra: RegraReclass) -> Optional[str]:
        """'' = sem filhos de vínculo (casa pelo código); conta = casou;
        None = tem filhos de vínculo mas nenhum casa."""
        reg_pb, reg_cta = PARES[pai]
        alvo_reg = reg_pb if regra.por_parte_b else reg_cta
        alvo_val = normaliza_conta(
            regra.conta_b if regra.por_parte_b else regra.relacionamento)

        tem_vinculo = False
        for j in self._filhos(regs, i, pai):
            f = regs[j]
            if f.reg in (reg_pb, reg_cta):
                tem_vinculo = True
            if f.reg == alvo_reg and alvo_val \
                    and normaliza_conta(f.campo(CAMPO_CONTA)) == alvo_val:
                return f.campo(CAMPO_CONTA)
        return "" if not tem_vinculo else None

    # ------------------------------------------------------------------ #
    def previa(self, arquivo: ArquivoECF,
               regras: Sequence[RegraReclass]) -> ResultadoReclass:
        res = ResultadoReclass()
        regs = arquivo.registros
        segmentos = self._segmentos(regs)
        com_valores = any(r.tem_valores for r in regras)

        if not self.tabela.disponivel:
            res.avisos.append(
                "Tabela oficial de códigos indisponível — os códigos destino "
                "não foram validados.")
        else:
            for r in regras:
                for pai in ("M300", "M350"):
                    if r.aplica_em(pai) and not self.tabela.existe(pai, r.para):
                        res.avisos.append(
                            f"Linha {r.linha}: o código destino {r.para} não "
                            f"consta na tabela oficial do {pai} — confira.")

        # ---- 1) candidatos por regra ---- #
        cand: Dict[int, List[_Candidato]] = {}
        for n, regra in enumerate(regras):
            achados: List[_Candidato] = []
            for per, ini, fim in segmentos:
                if com_valores and regra.tem_valores \
                        and regra.montante(per) is None:
                    continue
                for pai in PARES:
                    if not regra.aplica_em(pai):
                        continue
                    locais: List[_Candidato] = []
                    for i in range(ini, fim):
                        r = regs[i]
                        if r.reg != pai or r.campo(CAMPO_CODIGO) != regra.de:
                            continue
                        v = self._vinculo(regs, i, pai, regra)
                        if v is None:
                            continue
                        crit = ("vínculo" if v
                                else "código (lançamento sem vínculos)")
                        locais.append(_Candidato(i, pai, per, v, crit, 1.0))

                    if len(locais) > 1:
                        escolhido = self._desempata(regs, regra, locais, res)
                        if escolhido is None:
                            continue        # conflito já registrado
                        locais = [escolhido]
                    achados.extend(locais)
            cand[n] = achados
            if not achados:
                res.sem_alvo.append(
                    (regra, self._explica_sem_alvo(regs, regra)))

        if res.conflitos:
            return res

        # ---- 2) sem valores: comportamento clássico (troca integral) ---- #
        if not com_valores:
            self._fluxo_integral_classico(res, regras, regs, cand)
            res.correspondencias.sort(key=lambda c: c.indice)
            return res

        # ---- 3) com valores: saques por lançamento ---- #
        # localiza o lançamento de DESTINO já existente em cada período
        # (a ECF do PVA traz todos os códigos da tabela, muitos zerados)
        dest_idx: Dict[Tuple[str, str, str], int] = {}
        for per, ini, fim in segmentos:
            for i in range(ini, fim):
                r = regs[i]
                if r.reg in PARES:
                    dest_idx.setdefault((per, r.reg, r.campo(CAMPO_CODIGO)), i)

        por_idx: Dict[int, List[Tuple[RegraReclass, _Candidato]]] = {}
        for n, achados in cand.items():
            for c in achados:
                por_idx.setdefault(c.idx, []).append((regras[n], c))

        saldo_dest: Dict[int, float] = {}       # acumula somas nos destinos

        for i in sorted(por_idx):
            saques = sorted(por_idx[i], key=lambda rc: rc[0].linha)
            r = regs[i]
            per = saques[0][1].periodo
            valor = parse_valor_ecf(r.campo(CAMPO_VALOR))
            if valor is None:
                res.problemas.append(
                    f"{r.reg} {r.campo(CAMPO_CODIGO)} ({per}): lançamento "
                    f"sem VALOR numérico — período não alterado.")
                continue

            montantes = [(regra, c, regra.montante(c.periodo))
                         for regra, c in saques]
            montantes = [(rg, c, m) for rg, c, m in montantes if m]
            soma = sum(m for _, _, m in montantes)
            if soma > abs(valor) + TOL:
                det = "; ".join(f"linha {rg.linha}: R$ {fmt_brl(m)}"
                                for rg, _, m in montantes)
                msg = (f"{r.reg} {r.campo(CAMPO_CODIGO)} ({per}): a soma "
                       f"pedida (R$ {fmt_brl(soma)}) excede o valor do "
                       f"lançamento (R$ {fmt_brl(valor)}) — {det}. "
                       f"Período não alterado.")
                # detecção de De/Para invertidos: o "destino" tem o valor?
                rg0 = montantes[0][0]
                j = dest_idx.get((per, r.reg, rg0.para))
                if j is not None:
                    v_par = parse_valor_ecf(regs[j].campo(CAMPO_VALOR))
                    if v_par is not None and abs(abs(v_par) - soma) <= \
                            TOL * max(1, len(montantes)):
                        msg += (f" ⚠️ O lançamento {rg0.para} deste período "
                                f"tem exatamente R$ {fmt_brl(v_par)} — as "
                                f"colunas De (Y) e Para (Z) parecem estar "
                                f"INVERTIDAS no mapeamento.")
                res.problemas.append(msg)
                continue

            saldo = abs(valor)
            for k, (regra, c, m) in enumerate(montantes):
                ultima = (k == len(montantes) - 1)
                saldo_apos = saldo - m
                zera = ultima and abs(saldo_apos) <= TOL
                antes = self._linha_com_valor(r, saldo)
                j = dest_idx.get((per, c.pai, regra.para))
                if j is not None and j != i:
                    # -------- destino já existe: SOMA o montante nele ----- #
                    # (arquivo no padrão do PVA, com a tabela completa: a
                    # origem drenada permanece no arquivo, zerada)
                    rd = regs[j]
                    d_antes_v = saldo_dest.get(
                        j, abs(parse_valor_ecf(rd.campo(CAMPO_VALOR)) or 0.0))
                    d_depois_v = d_antes_v + m
                    saldo_dest[j] = d_depois_v
                    if zera:
                        copia = Registro(campos=list(r.campos))
                        copia.set_campo(CAMPO_VALOR, "0")
                        depois = copia.para_linha()
                    else:
                        depois = self._linha_com_valor(r, saldo_apos)
                    self._op_soma_destino(
                        res, regra, regs, i, c, m, antes, depois,
                        j, self._linha_com_valor(rd, d_antes_v),
                        self._linha_com_valor(rd, d_depois_v))
                elif zera and k == 0:
                    # destino não existe e o saque único drena → troca código
                    self._op_integral(res, regra, regs, i, c, m, antes)
                else:
                    depois = ("" if zera
                              else self._linha_com_valor(r, saldo_apos))
                    self._op_parcial(res, regra, regs, i, c, m,
                                     antes, depois)
                saldo = saldo_apos

        res.correspondencias.sort(key=lambda c: (c.indice, c.regra.linha))
        return res

    # ------------------------------------------------------------------ #
    def _desempata(self, regs, regra, locais, res) -> Optional[_Candidato]:
        pont = []
        for c in locais:
            sim = similaridade(regra.descricao, regs[c.idx].campo(CAMPO_DESC))
            m = regra.montante(c.periodo)
            v = parse_valor_ecf(regs[c.idx].campo(CAMPO_VALOR))
            bate = (m is not None and v is not None
                    and abs(m - abs(v)) <= TOL)
            pont.append((sim + (0.25 if bate else 0), sim, bate, c))
        pont.sort(key=lambda p: -p[0])
        top, seg = pont[0], pont[1]
        if (top[1] < LIMIAR_DESEMPATE and not top[2]) or \
                abs(top[0] - seg[0]) < 1e-9:
            res.conflitos.append(
                f"Linha {regra.linha}: há {len(locais)} lançamentos "
                f"{locais[0].pai} {regra.de} no período {locais[0].periodo} "
                f"com o mesmo vínculo, e a descrição não decide qual usar. "
                f"Diferencie as linhas da planilha e gere a prévia de novo.")
            return None
        c = top[3]
        c.criterio = f"{c.criterio} + descrição ({top[1]:.0%})"
        if top[2]:
            c.criterio += " + valor"
        c.score = top[1]
        return c

    def _fluxo_integral_classico(self, res, regras, regs, cand) -> None:
        dono: Dict[int, Tuple[int, _Candidato]] = {}
        for n, achados in cand.items():
            for c in achados:
                atual = dono.get(c.idx)
                if atual is None:
                    dono[c.idx] = (n, c)
                elif regras[atual[0]].para != regras[n].para:
                    sim_a = similaridade(regras[atual[0]].descricao,
                                         regs[c.idx].campo(CAMPO_DESC))
                    sim_n = similaridade(regras[n].descricao,
                                         regs[c.idx].campo(CAMPO_DESC))
                    melhor, sim = ((n, sim_n) if sim_n > sim_a
                                   else (atual[0], sim_a))
                    if sim < LIMIAR_DESEMPATE or abs(sim_n - sim_a) < 1e-9:
                        res.conflitos.append(
                            f"Linhas {regras[atual[0]].linha} e "
                            f"{regras[n].linha} disputam o mesmo lançamento "
                            f"({regs[c.idx].reg} {regras[n].de}) com destinos "
                            f"diferentes e a descrição não decide.")
                        continue
                    c2 = c if melhor == n else atual[1]
                    c2.criterio = f"vínculo + descrição ({sim:.0%})"
                    c2.score = sim
                    dono[c.idx] = (melhor, c2)
        if res.conflitos:
            return
        for i, (n, c) in sorted(dono.items()):
            r = regs[i]
            valor = abs(parse_valor_ecf(r.campo(CAMPO_VALOR)) or 0)
            self._op_integral(res, regras[n], regs, i, c, valor,
                              r.para_linha())

    # ------------------------------------------------------------------ #
    @staticmethod
    def _linha_com_valor(r: Registro, v: float) -> str:
        copia = Registro(campos=list(r.campos))
        copia.set_campo(CAMPO_VALOR, fmt_valor_ecf(v))
        return copia.para_linha()

    def _descricao_destino(self, pai: str, regra: RegraReclass,
                           atual: str) -> str:
        oficial = self.tabela.descricao(pai, regra.para)
        if self.alterar_descricao:
            return oficial or regra.descricao or atual
        return atual

    def _op_integral(self, res, regra, regs, i, c: _Candidato,
                     montante, antes) -> None:
        r = regs[i]
        copia = Registro(campos=Registro.de_linha(antes).campos)
        copia.set_campo(CAMPO_CODIGO, regra.para)
        copia.set_campo(CAMPO_DESC, self._descricao_destino(
            c.pai, regra, r.campo(CAMPO_DESC)))
        res.correspondencias.append(Correspondencia(
            regra=regra, indice=i, registro=c.pai, periodo=c.periodo,
            montante=montante, modo="integral",
            antes=antes, depois=copia.para_linha(),
            conta_casada=c.conta, criterio=c.criterio, score=c.score))

    def _op_parcial(self, res, regra, regs, i, c: _Candidato,
                    montante, antes, depois) -> None:
        r = regs[i]
        oficial = self.tabela.descricao(c.pai, regra.para)
        novo = Registro(campos=[
            c.pai, regra.para,
            oficial or regra.descricao or r.campo(CAMPO_DESC),
            r.campo(CAMPO_TIPO), r.campo(CAMPO_REL),
            fmt_valor_ecf(montante), ""])
        res.correspondencias.append(Correspondencia(
            regra=regra, indice=i, registro=c.pai, periodo=c.periodo,
            montante=montante, modo="parcial",
            antes=antes, depois=depois, novo=novo.para_linha(),
            conta_casada=c.conta, criterio=c.criterio, score=c.score))

    def _op_soma_destino(self, res, regra, regs, i, c: _Candidato,
                         montante, antes, depois, j,
                         destino_antes, destino_depois) -> None:
        res.correspondencias.append(Correspondencia(
            regra=regra, indice=i, registro=c.pai, periodo=c.periodo,
            montante=montante, modo="soma_destino",
            antes=antes, depois=depois,
            destino_indice=j, destino_antes=destino_antes,
            destino_depois=destino_depois,
            conta_casada=c.conta, criterio=c.criterio, score=c.score))

    # ------------------------------------------------------------------ #
    def _explica_sem_alvo(self, regs: List[Registro],
                          regra: RegraReclass) -> str:
        tot = {p: 0 for p in PARES}
        contas_pb: set = set()
        contas_cta: set = set()
        sem_filhos = 0
        for pai in PARES:
            reg_pb, reg_cta = PARES[pai]
            for i, r in enumerate(regs):
                if r.reg != pai or r.campo(CAMPO_CODIGO) != regra.de:
                    continue
                tot[pai] += 1
                filhos = self._filhos(regs, i, pai)
                achou_vinculo = False
                for j in filhos:
                    f = regs[j]
                    if f.reg == reg_pb:
                        contas_pb.add(f.campo(CAMPO_CONTA))
                        achou_vinculo = True
                    elif f.reg == reg_cta:
                        contas_cta.add(f.campo(CAMPO_CONTA))
                        achou_vinculo = True
                if not achou_vinculo:
                    sem_filhos += 1

        if tot["M300"] + tot["M350"] == 0:
            return (f"o código {regra.de} não aparece em nenhum M300/M350 "
                    f"deste arquivo")

        base = f"{tot['M300']} M300 e {tot['M350']} M350 com código {regra.de}"
        alvo = regra.conta_b if regra.por_parte_b else regra.relacionamento
        achadas = sorted(contas_pb if regra.por_parte_b else contas_cta)
        tipo = "Parte B" if regra.por_parte_b else "conta contábil"
        if achadas:
            return (f"{base}, mas nenhum vinculado à {tipo} '{alvo}'. "
                    f"{tipo.capitalize()}s presentes nesses lançamentos: "
                    f"{', '.join(achadas[:8])}"
                    f"{'…' if len(achadas) > 8 else ''}")
        return (f"{base}, porém sem nenhum filho de vínculo "
                f"({'M305/M355' if regra.por_parte_b else 'M310/M360'}) e "
                f"sem valores na planilha para transferir — informe os "
                f"valores mensais/anual para casar pelo código")

    # ------------------------------------------------------------------ #
    # Aplicação
    # ------------------------------------------------------------------ #
    def aplicar(self, arquivo: ArquivoECF, res: ResultadoReclass) -> int:
        if res.conflitos:
            raise ValueError("Há conflitos não resolvidos; nada foi aplicado.")
        regs = arquivo.registros

        por_origem: Dict[int, List[Correspondencia]] = {}
        for c in res.correspondencias:
            por_origem.setdefault(c.indice, []).append(c)

        # validação prévia: origem e destinos ainda são o que a prévia viu
        for i, ops in por_origem.items():
            ops.sort(key=lambda c: c.regra.linha)
            if regs[i].reg != ops[0].registro or \
                    regs[i].para_linha() != ops[0].antes:
                raise ValueError(
                    "O arquivo foi alterado após a prévia. Gere a prévia "
                    "novamente antes de aplicar.")
            for c in ops:
                if c.destino_indice is not None:
                    rd = regs[c.destino_indice]
                    if rd.reg != c.registro or \
                            rd.campo(CAMPO_CODIGO) != c.regra.para:
                        raise ValueError(
                            "O arquivo foi alterado após a prévia (destino). "
                            "Gere a prévia novamente antes de aplicar.")

        # plano de reconstrução
        replace: Dict[int, List[str]] = {}
        skip: set = set()
        insert_after: Dict[int, List[Registro]] = {}
        n_ops = 0

        for i, ops in por_origem.items():
            pai = ops[0].registro
            fim_origem = i
            for j in self._filhos(regs, i, pai):
                fim_origem = j

            for c in ops:
                n_ops += 1
                if c.modo == "integral":
                    replace[i] = Registro.de_linha(c.depois).campos
                    if self.alterar_descricao is False:
                        pass
                    continue

                # atualiza (ou remove) a origem
                if c.depois:
                    replace[i] = Registro.de_linha(c.depois).campos
                else:
                    skip.add(i)

                filhos_mov = self._planeja_vinculo(regs, i, pai, c,
                                                   replace, skip)

                if c.modo == "soma_destino":
                    j = c.destino_indice
                    replace[j] = Registro.de_linha(c.destino_depois).campos
                    if filhos_mov:
                        alvo = j
                        for k in self._filhos(regs, j, pai):
                            alvo = k
                        insert_after.setdefault(alvo, []).extend(filhos_mov)
                else:                                  # parcial → cria novo
                    novo = Registro(campos=Registro.de_linha(c.novo).campos)
                    insert_after.setdefault(fim_origem, []).append(novo)
                    insert_after[fim_origem].extend(filhos_mov)

            # origem removida não pode deixar filhos órfãos
            if i in skip:
                orfaos = [j for j in self._filhos(regs, i, pai)
                          if j not in skip]
                if orfaos:
                    skip.discard(i)
                    campos = Registro.de_linha(ops[-1].antes).campos
                    campos[CAMPO_VALOR] = "0,00"
                    replace[i] = campos

        # reconstrução em uma passada
        nova: List[Registro] = []
        for k, r in enumerate(regs):
            if k in replace:
                r.campos = replace[k]
            if k not in skip:
                nova.append(r)
            for extra in insert_after.get(k, []):
                nova.append(extra)
        arquivo.registros = nova
        return n_ops

    def _planeja_vinculo(self, regs: List[Registro], i: int, pai: str,
                         c: Correspondencia, replace: Dict[int, List[str]],
                         skip: set) -> List[Registro]:
        """Planeja a migração do vínculo; devolve os filhos a inserir."""
        if not c.conta_casada:
            return []
        reg_pb, reg_cta = PARES[pai]
        alvo = normaliza_conta(c.conta_casada)
        for j in self._filhos(regs, i, pai):
            if j in skip:
                continue
            f = regs[j]
            campos_atuais = replace.get(j, f.campos)
            if c.regra.por_parte_b and f.reg == reg_pb \
                    and normaliza_conta(campos_atuais[CAMPO_CONTA]) == alvo:
                vl = parse_valor_ecf(campos_atuais[CAMPO_VL_CTA]) or 0.0
                if abs(vl - c.montante) <= TOL:
                    skip.add(j)
                    return [Registro(campos=list(campos_atuais))]
                novos_campos = list(campos_atuais)
                novos_campos[CAMPO_VL_CTA] = fmt_valor_ecf(vl - c.montante)
                replace[j] = novos_campos
                return [Registro(campos=[
                    reg_pb, campos_atuais[CAMPO_CONTA],
                    fmt_valor_ecf(c.montante), campos_atuais[CAMPO_IND]])]
            if not c.regra.por_parte_b and f.reg == reg_cta \
                    and normaliza_conta(campos_atuais[CAMPO_CONTA]) == alvo:
                skip.add(j)
                return [Registro(campos=list(campos_atuais))]
        return []
