"""Reclassificação de códigos M300/M350 a partir da planilha De→Para."""
from __future__ import annotations

import pytest

from editor_ecf.core import (
    LeitorReclass, ParserECF, Reclassificador, RegraReclass,
)

ECF = [
    "|0000|LECF|0012|02139940000191|EMPRESA TESTE|01012025|31122025|0|0|N|N|N||",
    "|0001|0|", "|0990|3|",
    "|M001|0|",
    "|M300|84.05|CUSTO DE CONSTRUCAO|A|4|65184,88||",   # ligado a conta contábil
    "|M310|6105219999990|",
    "|M300|84.05|DIFERIMENTO AVP|A|1|8398088,01||",     # ligado à Parte B 121
    "|M305|121|8398088,01|C|",
    "|M990|5|",
    "|M350|84.05|CUSTO DE CONSTRUCAO|A|4|65184,88||",   # espelho na CSLL
    "|M360|6105219999990|",
    "|9001|0|", "|9990|0|", "|9999|0|",
]


@pytest.fixture()
def arq(tmp_path):
    p = tmp_path / "ecf.txt"
    p.write_bytes(("\r\n".join(ECF) + "\r\n").encode("latin-1"))
    return ParserECF().ler(str(p))


def _regra(de, para, rel="", cb="", linha=2):
    return RegraReclass(linha=linha, de=de, para=para,
                        relacionamento=rel, conta_b=cb)


def test_reclassifica_por_conta_contabil(arq):
    regras = [_regra("84.05", "82.05", rel="6105219999990", cb="121")]
    res = Reclassificador().previa(arq, regras)
    # casa o M300 ligado à conta e o espelho M350 — não o da Parte B
    assert len(res.correspondencias) == 2
    assert {c.registro for c in res.correspondencias} == {"M300", "M350"}
    assert all("82.05" in c.depois for c in res.correspondencias)


def test_reclassifica_por_parte_b(arq):
    regras = [_regra("84.05", "82.10", rel="PARTE B", cb="121")]
    res = Reclassificador().previa(arq, regras)
    assert len(res.correspondencias) == 1
    c = res.correspondencias[0]
    assert "DIFERIMENTO AVP" in c.antes and "82.10" in c.depois


def test_parte_b_nao_afeta_lancamento_de_conta(arq):
    """A regra de Parte B não pode reclassificar o vínculo por conta."""
    regras = [_regra("84.05", "82.10", rel="PARTE B", cb="121")]
    res = Reclassificador().previa(arq, regras)
    Reclassificador().aplicar(arq, res)
    codigos = [r.campo(1) for r in arq.registros if r.reg == "M300"]
    assert "84.05" in codigos      # o ligado à conta contábil ficou intacto
    assert "82.10" in codigos


def test_conta_errada_nao_casa(arq):
    regras = [_regra("84.05", "82.05", rel="9999999999", cb="121")]
    res = Reclassificador().previa(arq, regras)
    assert not res.correspondencias
    assert len(res.sem_alvo) == 1


def test_conflito_bloqueia_aplicacao(arq):
    regras = [_regra("84.05", "82.10", rel="PARTE B", cb="121", linha=3),
              _regra("84.05", "82.05", rel="PARTE B", cb="121", linha=4)]
    res = Reclassificador().previa(arq, regras)
    assert res.conflitos and not res.aplicavel
    with pytest.raises(ValueError):
        Reclassificador().aplicar(arq, res)


def test_descricao_preservada_por_padrao(arq):
    regras = [_regra("84.05", "82.05", rel="6105219999990", cb="121")]
    regras[0].descricao = "OUTRA DESCRICAO"
    res = Reclassificador().previa(arq, regras)
    Reclassificador().aplicar(arq, res)
    m300 = [r for r in arq.registros if r.reg == "M300"][0]
    assert m300.campo(2) == "CUSTO DE CONSTRUCAO"


def test_leitor_planilha(tmp_path):
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws["Y1"], ws["Z1"] = "LMTE", "EY"       # cabeçalho deve ser ignorado
    ws["B2"], ws["C2"] = "84.05", "CUSTO"
    ws["F2"], ws["G2"] = "PARTE B", "121"
    ws["Y2"], ws["Z2"] = "84.05", "82.10"
    ws["Y3"], ws["Z3"] = "92", "92"          # igual → ignorada com aviso
    p = tmp_path / "rec.xlsx"
    wb.save(p)

    regras, avisos = LeitorReclass().ler(str(p))
    assert len(regras) == 1
    assert regras[0].de == "84.05" and regras[0].para == "82.10"
    assert regras[0].por_parte_b
    assert avisos            # avisou sobre a linha com De == Para


# ----------------------- desempate por descrição ------------------------- #
ECF_AMBIGUO = [
    "|0000|LECF|0012|02139940000191|X|01012025|31122025|0|0|N|N|N||",
    "|0001|0|", "|0990|3|", "|M001|0|",
    "|M300|84.05|Diferimento AVP Art. 84 Inciso II|A|1|90353050,93||",
    "|M305|121|90353050,93|C|",
    "|M300|84.05|Provisao PIS/COFINS - ativo de contrato|A|1|4033459,27||",
    "|M305|121|4033459,27|C|",
    "|M990|7|", "|9001|0|", "|9990|0|", "|9999|0|",
]


@pytest.fixture()
def arq_ambiguo(tmp_path):
    p = tmp_path / "amb.txt"
    p.write_bytes(("\r\n".join(ECF_AMBIGUO) + "\r\n").encode("latin-1"))
    return ParserECF().ler(str(p))


def test_desempata_pela_descricao(arq_ambiguo):
    """Mesmo código + mesma Parte B: a descrição decide qual é qual."""
    regras = [
        RegraReclass(linha=7, de="84.05", para="82.10",
                     relacionamento="PARTE B", conta_b="121",
                     descricao="Diferimento AVP Art. 84, Inciso II"),
        RegraReclass(linha=8, de="84.05", para="82.05",
                     relacionamento="PARTE B", conta_b="121",
                     descricao="PROVISÃO PIS/COFINS - ATIVO DE CONTRATO"),
    ]
    res = Reclassificador().previa(arq_ambiguo, regras)
    assert not res.conflitos
    assert len(res.correspondencias) == 2
    por_desc = {c.regra.linha: c for c in res.correspondencias}
    assert "Diferimento AVP" in por_desc[7].antes
    assert "82.10" in por_desc[7].depois
    assert "PIS/COFINS" in por_desc[8].antes
    assert "82.05" in por_desc[8].depois


def test_descricoes_irreconheciveis_geram_conflito(arq_ambiguo):
    regras = [
        RegraReclass(linha=7, de="84.05", para="82.10",
                     relacionamento="PARTE B", conta_b="121",
                     descricao="ZZZZ QQQQ"),
        RegraReclass(linha=8, de="84.05", para="82.05",
                     relacionamento="PARTE B", conta_b="121",
                     descricao="WWWW KKKK"),
    ]
    res = Reclassificador().previa(arq_ambiguo, regras)
    assert res.conflitos and not res.aplicavel


def test_le_planilha_real_do_usuario():
    """A planilha real deve render 8 regras, cabeçalho descartado."""
    import os
    caminho = "/mnt/user-data/uploads/Reclassificação.xlsx"
    if not os.path.exists(caminho):
        pytest.skip("planilha real indisponível")
    regras, _ = LeitorReclass().ler(caminho)
    assert len(regras) == 8
    l7 = [r for r in regras if r.linha == 7][0]
    assert l7.por_parte_b and l7.conta_b == "121" and l7.para == "82.10"
    l6 = [r for r in regras if r.linha == 6][0]
    assert not l6.por_parte_b and l6.relacionamento == "6105219999990"


# ------------------- transferência de valores por período ------------------ #
ECF_PERIODOS = [
    "|0000|LECF|0012|02139940000191|X|01012025|31122025|0|0|N|N|N||",
    "|0001|0|", "|0990|3|", "|M001|0|",
    "|M030|01012025|31012025|A01|",
    "|M300|84.05|CPC 47 Ajustes|A|4|100000,00||",
    "|M310|6105219999990|",
    "|M310|7777777777777|",                       # conta que não migra
    "|M030|01012025|31122025|A00|",
    "|M300|84.05|CPC 47 Ajustes|A|4|500000,00||",
    "|M310|6105219999990|",
    "|M990|7|", "|9001|0|", "|9990|0|", "|9999|0|",
]


@pytest.fixture()
def arq_per(tmp_path):
    p = tmp_path / "per.txt"
    p.write_bytes(("\r\n".join(ECF_PERIODOS) + "\r\n").encode("latin-1"))
    return ParserECF().ler(str(p))


def _regra_valores(jan=None, anual=None, **kw):
    vals = [None] * 13
    vals[0], vals[12] = jan, anual
    return RegraReclass(linha=6, de="84.05", para="82.05",
                        relacionamento="6105219999990", conta_b="121",
                        descricao="CPC 47", valores=vals, **kw)


def test_transferencia_parcial_move_valor_e_conta(arq_per):
    """Linha 6 do usuário: tira 65.184,88 do 84.05 e leva ao 82.05."""
    regra = _regra_valores(jan=65184.88, anual=500000.0)
    rec = Reclassificador()
    res = rec.previa(arq_per, [regra])
    assert res.aplicavel and not res.problemas
    modos = {(c.periodo, c.modo) for c in res.correspondencias}
    assert ("A01", "parcial") in modos       # 100.000 > 65.184,88
    assert ("A00", "integral") in modos      # 500.000 == 500.000

    rec.aplicar(arq_per, res)
    linhas = [r.para_linha() for r in arq_per.registros]
    # A01: De reduzido, novo criado com a conta migrada
    assert "|M300|84.05|CPC 47 Ajustes|A|4|34815,12||" in linhas
    i_novo = next(i for i, l in enumerate(linhas) if l.startswith("|M300|82.05|"))
    assert linhas[i_novo + 1] == "|M310|6105219999990|"
    # a outra conta permaneceu no De
    i_de = next(i for i, l in enumerate(linhas) if "34815,12" in l)
    assert linhas[i_de + 1] == "|M310|7777777777777|"


def test_periodo_sem_valor_e_pulado(arq_per):
    regra = _regra_valores(jan=None, anual=500000.0)   # só o anual
    res = Reclassificador().previa(arq_per, [regra])
    periodos = {c.periodo for c in res.correspondencias}
    assert periodos == {"A00"}


def test_montante_maior_que_lancamento_vira_problema(arq_per):
    regra = _regra_valores(jan=999999.0, anual=500000.0)
    res = Reclassificador().previa(arq_per, [regra])
    assert any("A01" in p for p in res.problemas)
    # o anual continua aplicável
    assert any(c.periodo == "A00" for c in res.correspondencias)


def test_split_parcial_da_parte_b(tmp_path):
    linhas = [
        "|0000|LECF|0012|02139940000191|X|01012025|31122025|0|0|N|N|N||",
        "|0001|0|", "|0990|3|", "|M001|0|",
        "|M030|01012025|31122025|A00|",
        "|M300|84.05|Diferimento AVP|A|1|100000,00||",
        "|M305|121|100000,00|C|",
        "|M990|4|", "|9001|0|", "|9990|0|", "|9999|0|",
    ]
    p = tmp_path / "pb.txt"
    p.write_bytes(("\r\n".join(linhas) + "\r\n").encode("latin-1"))
    arq = ParserECF().ler(str(p))

    vals = [None] * 13
    vals[12] = 30000.0
    regra = RegraReclass(linha=7, de="84.05", para="82.10",
                         relacionamento="PARTE B", conta_b="121",
                         descricao="Diferimento AVP", valores=vals)
    rec = Reclassificador()
    res = rec.previa(arq, [regra])
    assert res.correspondencias[0].modo == "parcial"
    rec.aplicar(arq, res)
    ls = [r.para_linha() for r in arq.registros]
    assert "|M305|121|70000,00|C|" in ls      # M305 original reduzido
    assert "|M305|121|30000,00|C|" in ls      # M305 novo sob o 82.10
    i_novo = next(i for i, l in enumerate(ls) if l.startswith("|M300|82.10|"))
    assert ls[i_novo + 1] == "|M305|121|30000,00|C|"


def test_totalizadores_apos_transferencia(arq_per, layout, versao):
    """As inserções do modo parcial devem manter o arquivo autoconsistente."""
    from editor_ecf.core import Contador
    regra = _regra_valores(jan=65184.88, anual=500000.0)
    rec = Reclassificador()
    res = rec.previa(arq_per, [regra])
    rec.aplicar(arq_per, res)
    Contador().recalcular(arq_per, layout.ordem_registros(versao))
    m = {r.reg: r.campos for r in arq_per.registros}
    assert m["9999"][1] == str(arq_per.total_linhas)
    assert m["M990"][1] == str(sum(1 for r in arq_per.registros
                                   if r.bloco == "M"))


def test_codigo_destino_inexistente_gera_aviso(arq_per):
    regra = _regra_valores(jan=None, anual=500000.0)
    regra.para = "99999.99"
    res = Reclassificador().previa(arq_per, [regra])
    assert any("99999.99" in a for a in res.avisos)


# ---------------- saques múltiplos de lançamento agregado ------------------ #
ECF_AGREGADO = [
    "|0000|LECF|0012|02139940000191|X|01012025|31122025|0|0|N|N|N||",
    "|0001|0|", "|0990|3|", "|M001|0|",
    "|M030|01012025|31122025|A00|",
    "|M300|84.05|CPC 47 Ajustes|A|4|100000,00||",     # agregado, SEM filhos
    "|M990|3|", "|9001|0|", "|9990|0|", "|9999|0|",
]


@pytest.fixture()
def arq_agregado(tmp_path):
    p = tmp_path / "agg.txt"
    p.write_bytes(("\r\n".join(ECF_AGREGADO) + "\r\n").encode("latin-1"))
    return ParserECF().ler(str(p))


def _regra_anual(linha, para, anual, rel="6105219999990", cb="121",
                 desc="CPC 47"):
    vals = [None] * 13
    vals[12] = anual
    return RegraReclass(linha=linha, de="84.05", para=para,
                        relacionamento=rel, conta_b=cb, descricao=desc,
                        valores=vals)


def test_lancamento_sem_filhos_casa_pelo_codigo(arq_agregado):
    res = Reclassificador().previa(arq_agregado,
                                   [_regra_anual(6, "82.05", 30000.0)])
    assert len(res.correspondencias) == 1
    assert "código" in res.correspondencias[0].criterio


def test_varias_linhas_sacam_do_mesmo_lancamento(arq_agregado):
    regras = [_regra_anual(6, "82.05", 30000.0),
              _regra_anual(7, "82.10", 50000.0, rel="PARTE B"),
              _regra_anual(8, "82.05", 20000.0, rel="PARTE B")]
    rec = Reclassificador()
    res = rec.previa(arq_agregado, regras)
    assert not res.problemas and not res.conflitos
    assert len(res.correspondencias) == 3
    # a última sacada zera o De -> removido
    assert res.correspondencias[-1].depois == ""

    rec.aplicar(arq_agregado, res)
    ls = [r.para_linha() for r in arq_agregado.registros]
    assert not any(l.startswith("|M300|84.05|") for l in ls)   # De removido
    assert sum(1 for l in ls if l.startswith("|M300|82.05|")) == 2
    assert sum(1 for l in ls if l.startswith("|M300|82.10|")) == 1


def test_soma_dos_saques_excede_gera_problema(arq_agregado):
    regras = [_regra_anual(6, "82.05", 80000.0),
              _regra_anual(7, "82.10", 50000.0, rel="PARTE B")]
    res = Reclassificador().previa(arq_agregado, regras)
    assert not res.correspondencias
    assert any("excede" in p for p in res.problemas)


def test_diagnostico_conta_divergente(arq):
    """Regra com conta que não existe: o diagnóstico lista as presentes."""
    regra = _regra(  # helper antigo: sem valores
        "84.05", "82.05", rel="1234567890", cb="121")
    res = Reclassificador().previa(arq, [regra])
    assert res.sem_alvo
    _, detalhe = res.sem_alvo[0]
    assert "6105219999990" in detalhe    # mostra a conta que existe de fato


def test_conta_com_mascara_casa(tmp_path):
    """'6.105.21.9999990' na ECF casa com '6105219999990' da planilha."""
    linhas = [
        "|0000|LECF|0012|02139940000191|X|01012025|31122025|0|0|N|N|N||",
        "|0001|0|", "|0990|3|", "|M001|0|",
        "|M300|84.05|CPC 47|A|4|1000,00||",
        "|M310|6.105.21.9999990|",
        "|M990|3|", "|9001|0|", "|9990|0|", "|9999|0|",
    ]
    p = tmp_path / "m.txt"
    p.write_bytes(("\r\n".join(linhas) + "\r\n").encode("latin-1"))
    arq2 = ParserECF().ler(str(p))
    regra = _regra("84.05", "82.05", rel="6105219999990", cb="121")
    res = Reclassificador().previa(arq2, [regra])
    assert len(res.correspondencias) == 1


# --------------- destino pré-existente (padrão PVA) ------------------------ #
ECF_PVA = [
    "|0000|LECF|0012|02139940000191|X|01012025|31122025|0|0|N|N|N||",
    "|0001|0|", "|0990|3|", "|M001|0|",
    "|M030|01012025|31122025|A00|",
    "|M300|82.05|Contratos concessao receb|A|4|0||",     # destino zerado
    "|M300|84.05|CPC 47 Ajustes|A|4|100000,00||",        # origem com valor
    "|M990|4|", "|9001|0|", "|9990|0|", "|9999|0|",
]


@pytest.fixture()
def arq_pva(tmp_path):
    p = tmp_path / "pva.txt"
    p.write_bytes(("\r\n".join(ECF_PVA) + "\r\n").encode("latin-1"))
    return ParserECF().ler(str(p))


def test_soma_no_destino_existente_sem_duplicar(arq_pva):
    regras = [_regra_anual(6, "82.05", 60000.0),
              _regra_anual(8, "82.05", 40000.0, rel="PARTE B")]
    rec = Reclassificador()
    res = rec.previa(arq_pva, regras)
    assert all(c.modo == "soma_destino" for c in res.correspondencias)
    rec.aplicar(arq_pva, res)
    ls = [r.para_linha() for r in arq_pva.registros]
    # um único 82.05, com a soma; origem mantida zerada (padrão PVA)
    l8205 = [l for l in ls if l.startswith("|M300|82.05|")]
    assert len(l8205) == 1 and "100000,00" in l8205[0]
    l8405 = [l for l in ls if l.startswith("|M300|84.05|")]
    assert len(l8405) == 1 and "|0||" in l8405[0]


def test_deteccao_de_para_invertidos(arq_pva):
    """De=82.05 (zerado) e Para=84.05 (com o valor) → aviso de inversão."""
    vals = [None] * 13
    vals[12] = 100000.0
    regra = RegraReclass(linha=6, de="82.05", para="84.05",
                         relacionamento="6105219999990", conta_b="121",
                         descricao="X", valores=vals)
    res = Reclassificador().previa(arq_pva, [regra])
    assert res.problemas
    assert any("INVERTIDAS" in p for p in res.problemas)
